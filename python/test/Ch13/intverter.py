#!
#use updatedProduceSales.xlsx
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

#source stuff
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = 'Spam'

#iterate through each cell
for rowObj in sheet.rows:
    for cellObj in rowObj:
        sheet2.cell(row=cellObj.row,column=cellObj.column).value = sheet.cell(row=cellObj.column,column=cellObj.row).value
#save new sheet
wb2.save('moreProduceSales.xlsx')
print('done')