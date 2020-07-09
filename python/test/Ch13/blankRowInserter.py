#!
import openpyxl,sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

#getparts
#1 = N
#2 = M
#3 = file

if len(sys.argv) < 3:
    sys.exit('Need 2 ints and an excel file in the same folder as program')
n = int(sys.argv[1])
m = int(sys.argv[2])
exfile = sys.argv[3]
#open .xl
wb = openpyxl.load_workbook(exfile)
sheet = wb['Sheet']
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = 'Spam'

#for loop throught row
mr = sheet.max_row
mc = sheet.max_column
for rowNum in range(1,mr):
    for colNum in range(1,mc):
        cellVal = sheet.cell(row=rowNum, column=colNum).value
        # if cellVal == None:
        #     continue
        #if cell >= N add M to each
        if rowNum >= n:
            sheet2.cell(row=rowNum+ m, column=colNum).value = cellVal
        else:
            sheet2.cell(row=rowNum, column=colNum).value = cellVal

# for rowNum in range(1,sheet.max_row):
#     for colNum in range(1,sheet.max_column):
#         cellVal = sheet.cell(row=rowNum, column=colNum).value
#         # if cellVal == None:
#         #     continue
#         #if cell >= N add M to each
#         if rowNum >= n:
#             sheet.cell(row=rowNum+ m, column=colNum).value = cellVal
#         # else:
#         #     sheet.cell(row=rowNum, column=colNum).value = cellVal

#save new .xl
wb2.save('blanksMade.xlsx')
print('done')