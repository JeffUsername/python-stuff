#!
#use updatedProduceSales.xlsx
import openpyxl, re
from openpyxl.utils import get_column_letter, column_index_from_string

#open testToSheet.xlsx and get sheet info
wb = openpyxl.load_workbook('testToSheet.xlsx')
sheet = wb['Spam']

colCounter =1
#for columns
for colObj in sheet.columns:

    #create a text as writeable numbered by colObj
    newfile = 'txtFile' + str(colCounter) + '.txt'
    #open file
    bacon = open(newfile, 'w')
    #for cell in colObj.rows
    for cellObj in colObj:
        #get cell to string +\n
        line =  str(sheet.cell(row=cellObj.row,column=cellObj.column).value)
        line = othersRegex = re.sub('None','',line)
        bacon.write(line)
    #close file
    bacon.close()
    colCounter +=1
print('done')