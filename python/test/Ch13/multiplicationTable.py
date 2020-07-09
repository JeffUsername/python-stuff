#!
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
#make new .xl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'

fontObj1 = Font(name='Times New Roman', bold=True) #bold for the first row and col
#for loop throught row
for y in range(1,8):
    #for loop through columns @number
    for x in range(1,8):
        
        #if x and y =1 continue
        if x == 1 and y ==1:
            continue
        #elif x =1, cell = y-1 make bold
        elif x == 1:
            val =y-1
            sheet.cell(row=x, column=y).value = val
            loc = get_column_letter(x) + str(y)
            sheet[loc].font = fontObj1
        #elif y = 1, cell = x-1 make bold
        elif y == 1:
            val =x-1
            sheet.cell(row=x, column=y).value = val
            loc = get_column_letter(x) + str(y)
            sheet[loc].font = fontObj1
        #else: cell x[current][1].val-1*y[1][current].val-1
        else:
            valforX =sheet.cell(row=1, column=y).value
            valforY = val =sheet.cell(row=x, column=1).value
            sheet.cell(row=x, column=y).value = valforX*valforY

wb.save('idk.xlsx')
print('done')