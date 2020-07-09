#!
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
#get files
# i =input('enter file names seperated by a space: ')
# files = i.split()
files = ['t1.txt','t2.txt','t3.txt'] #i'm too lazt to type it each time

#make area fpr xlsx file.
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = 'Spam'
colcount = 1
#for list of files
for items in files:
    
    #open file at read line
    theFile = open(items) #open file
    reading = theFile.readlines() #read by lines
    rowcount = 1
    for line in reading:
    #for rows line numbers?
        sheet2.cell(row=rowcount,column=colcount).value = line
        #input line info to row at line#, column list#
        rowcount+=1
    colcount += 1
    #close file
    theFile.close()
#save file
wb2.save('testToSheet.xlsx')
print('done')